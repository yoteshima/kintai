# -*- coding: utf-8 -*-

from django.shortcuts import render, redirect, reverse
from django.http import HttpResponse
import calendar
from datetime import datetime, timedelta, time

from .forms import TimeStampForm

from .models import WorkReport, WorkDetail, Status, ReportSettings

from apps.account.models import User

import openpyxl
from openpyxl import utils
from openpyxl.styles import Protection
import email.utils

import collections

from . import time_util


def show_time_stamp(request):
    return render(request, 'time_stamp/time_stamp.html')

def show_work_report(request):
    """
        勤務報告書のリンクを押したときの処理
    """

    # リクエストクエリから年と月を取得します
    year = request.GET.get("year")
    month = request.GET.get("month")

    # リクエストクエリがない場合は当日
    if not year:
        year = datetime.now().year
    else:
        year = int(year)

    if not month:
        month = datetime.now().month
    else:
        month = int(month)

    user_id = request.user.no # ログインしたユーザの社員番号

    report_period = datetime(year=year, month=month, day=1) # 対象月の初日

    # 対象月の報告書
    # work_report = WorkReport.objects.get_or_create(
    work_report = get_work_report(user_id=user_id, report_period=report_period)

    # 当月の報告書明細
    work_details = get_work_detail(report_no=work_report.report_no,report_period=report_period,user_id=user_id)

    month_range = calendar.monthrange(year,month)[1]
    first_date = datetime.strptime(str(year)+str(month).zfill(2),"%Y%m")
    month_date = [first_date + timedelta(days=i) for i in range(month_range)]

    ret_dict = {
        'month_date' : month_date,
        "work_report" : work_report,
        "work_details" : work_details.values()
    }

    return render(request, 'work_report/work_report.html', ret_dict)

def show_settings(request):
    user_id = request.user.no # ログインしたユーザの社員番号

    #ユーザIDからReportSettingsを検索してレコードが存在しない場合は作成
    report_settings = ReportSettings.objects.get_or_create(
        user_id = User.objects.get(no=user_id)
    )[0]

    return render(request, 'settings/settings.html', {'report_settings': report_settings})


def exec_ajax(request):

    user_id = request.user.no

    if request.method == "GET":
        data = dict(request.GET.items())

        if data["type"] == "change_year_month":
            year = int(data["year"])
            month = int(data["month"])

            ret_dict = {
                "year":year,
                "month":month,
            }

            return change_report(request,ret_dict)


    elif request.method == "POST":
        data = dict(request.POST.items())

        # 打刻画面
        if data["type"] == "time_stamp":
            report_period = get_first_date()

            # 当月の報告書
            # work_report = WorkReport.objects.get_or_create(
            work_report = get_work_report(user_id=user_id, report_period=report_period)

            # 当月分の報告書明細
            work_details = get_work_detail(report_no=work_report.report_no,report_period=report_period,user_id=user_id)
            # 当日の報告書明細
            work_detail = work_details[datetime.today().date()]

            stamp_time = data["time"].replace(" ","")

            # 出勤
            if data["method"] == "entry-work":
                work_detail.start_time = stamp_time
                if not work_detail.break1 and not work_detail.break2: # どっちにも入っていない場合デフォルト値を設定する。
                    work_detail.break1 = 1
                work_detail.save()

            # 退勤
            elif data["method"] == "leave-work":
                work_detail.end_time = stamp_time
                work_detail.save()

            # 休憩入
            elif data["method"] == "entry-break":
                work_detail.start_break = stamp_time
                work_detail.save()

            # 休憩戻
            elif data["method"] == "leave-break":
                work_detail.end_break = stamp_time
                work_detail.save()

            return HttpResponse()

        # 作業報告書
        elif data["type"] == "change_work_report":
            # 入力値に誤りがなければresponseを返す

            try:
                target = data["target"]
                year = int(data["year"])
                month = int(data["month"])

                report_period = datetime(year=year, month=month, day=1) # 対象月の初日

                # 対象月の報告書
                # work_report = WorkReport.objects.get_or_create(
                work_report = get_work_report(user_id=user_id, report_period=report_period)

                if target == "project_name":
                    project_name = data["project_name"]
                    work_report.project_name = project_name
                    work_report.save()

                elif target == "site_work_time":
                    site_work_time = time_util.convert_to_clock_time(data["site_work_time"])

                    if not site_work_time:
                        raise Exception

                    work_report.site_work_time = site_work_time
                    work_report.save()

                elif target == "report_detail":
                    date=int(data["date"])
                    start_time=time_util.convert_to_clock_time(data["start_time"])
                    end_time=time_util.convert_to_clock_time(data["end_time"])
                    break1_time=time_util.convert_to_float_time(time_util.convert_to_clock_time(data["break1_time"]))
                    break2_time=time_util.convert_to_float_time(time_util.convert_to_clock_time(data["break2_time"]))
                    status_code=int(data["status_code"])
                    remarks=data["remarks"]

                    # 対象月分の報告書明細
                    work_details = get_work_detail(report_no=work_report.report_no,report_period=report_period,user_id=user_id)

                    # 対象日
                    target_date = report_period.replace(day=date)
                    # 対象日の報告書明細
                    work_detail = work_details[target_date]

                    work_detail.start_time = start_time
                    work_detail.end_time = end_time
                    work_detail.break1 = break1_time
                    work_detail.break2 = break2_time
                    work_detail.status = Status.objects.get(code=status_code)
                    work_detail.remarks = remarks
                    work_detail.save()

                elif target is None:
                    raise Exception

                return HttpResponse(status=200)

            except Exception as e:
                return HttpResponse(status=500) # なんか不備がある場合の戻り値

        # 設定値
        elif data["type"] == "change_report_settings":
            try:
                report_settings = ReportSettings.objects.get_or_create(
                    user_id = User.objects.get(no=user_id)
                )[0]

                start_time=time_util.convert_to_clock_time(data["start_time"])
                end_time=time_util.convert_to_clock_time(data["end_time"])
                break1_time=time_util.convert_to_clock_time(data["break1_time"])
                break2_time=time_util.convert_to_clock_time(data["break2_time"])
                holidays = data["holidays"]
                project_name = data["project_name"]

                report_settings.start_time = start_time
                report_settings.end_time = end_time
                report_settings.break1 = break1_time
                report_settings.break2 = break2_time
                report_settings.holidays = holidays
                report_settings.project_name = project_name
                report_settings.save()

                return HttpResponse(status=200)

            except Exception as e:
                return HttpResponse(status=500)


def get_work_report(report_period,user_id):
    """
        report_periodの一か月分で検索したwork_reportを返す。
        レコードが存在しない場合、作成してデフォルト値を適用して返す
    """
    try:
        work_report = WorkReport.objects.get(
            user_id = User.objects.get(no=user_id),
            report_period = report_period
        )

    except Exception as e: # レコードが存在しない場合
        # 設定値を取得
        report_settings = ReportSettings.objects.get_or_create(
            user_id = User.objects.get(no=user_id)
        )[0]

        work_report = WorkReport.objects.create(
            user_id = User.objects.get(no=user_id),
            report_period = report_period
        )
        work_report.project_name = report_settings.project_name
        work_report.save()

    return work_report


def get_work_detail(report_no,report_period,user_id):
    """
        report_no,report_periodの一か月分で検索したwork_detailと日付のディクショナリを返す。
        レコードが存在しない場合、作成してデフォルト値を適用して返す
    """
    report_year = report_period.year
    report_month = report_period.month

    month_length = calendar.monthrange(report_period.year,report_period.month)[1]
    month_dates = [report_period.replace(day=day+1) for day in range(month_length)]


    try:
        work_details = {work_date:WorkDetail.objects.get(
            report_no=WorkReport.objects.get(report_no=report_no),
            work_date=work_date
        ) for work_date in month_dates}

    except Exception as e: # レコードが存在しない場合
        # 設定値を取得
        report_settings = ReportSettings.objects.get_or_create(
            user_id = User.objects.get(no=user_id)
        )[0]
        holiday_week = report_settings.holiday_week

        # レコード作成
        work_details = {work_date:WorkDetail.objects.create(
            report_no=WorkReport.objects.get(report_no=report_no),
            work_date=work_date
        ) for work_date in month_dates}

        # 設定値を適用
        for work_date,work_detail in work_details.items():
            if not work_date.strftime("%a") in holiday_week:
                work_detail.start_time = report_settings.start_time
                work_detail.end_time = report_settings.end_time
                work_detail.break1 = time_util.convert_to_float_time(report_settings.break1)
                work_detail.break2 = time_util.convert_to_float_time(report_settings.break2)
                if work_date.strftime("%a") in ("Sun","Sat"):
                    work_detail.status = Status.objects.get(code=8) # シフト出勤
            else:
                if not work_date.strftime("%a") in ("Sun","Sat"):
                    work_detail.status = Status.objects.get(code=9) # シフト休

            work_detail.save()

    return work_details


def get_first_date():
    return datetime.today().date().replace(day=1)


def flatten(l):
    for el in l:
        if isinstance(el, collections.abc.Iterable) and not isinstance(el, (str, bytes)):
            yield from flatten(el)
        else:
            yield el

def reset_report(request):
    """
        勤務表をクリアして初期値を適用します
    """
    year = request.GET.get("year")
    month = request.GET.get("month")

    if year and month:

        user_id = request.user.no
        user_name = request.user.username

        report_period = datetime(year=int(year),month=int(month),day=1)


        # 対象月の報告書
        # work_report = WorkReport.objects.get_or_create(
        work_report = get_work_report(user_id=user_id, report_period=report_period)

        # クリア
        [work_detail.delete() for work_detail in get_work_detail(report_no=work_report.report_no,report_period=report_period,user_id=user_id).values()]
        # 対象月分の報告書明細
        # work_details = get_work_detail(report_no=work_report.report_no,report_period=report_period,user_id=user_id)
        work_report.delete()


        return show_work_report(request)




def export_excel(request):
    """
        Excelファイルを出力
    """

    # リクエストクエリから年と月を取得します
    year = request.GET.get("year")
    month = request.GET.get("month")

    if year and month:

        PASSWORD = "jobcrown1234"

        user_id = request.user.no
        user_name = request.user.username

        report_period = datetime(year=int(year),month=int(month),day=1)


        # 対象月の報告書
        # work_report = WorkReport.objects.get_or_create(
        work_report = get_work_report(user_id=user_id, report_period=report_period)
        # 対象月分の報告書明細
        work_details = get_work_detail(report_no=work_report.report_no,report_period=report_period,user_id=user_id)


        # 出力ファイル名
        # file_name = f"自社用作業報告書_{user_id}_{user_name}_{report_period.year}年{report_period.month}月_v106.xlsm"
        file_name = f"自社用作業報告書_{user_id}_{user_name}_{report_period.year}年{report_period.month}月_v106a.xlsm"

        # 作業報告書テンプレートを読み込み
        # wb = openpyxl.load_workbook("static/excel/自社用作業報告書_社員番号_名前_2019年X月_v106.xlsm", keep_vba=True)
        wb = openpyxl.load_workbook("static/excel/自社用作業報告書_社員番号_名前_2019年X月_v106a.xlsm", keep_vba=True)


        sheet = wb["実績表"]

        sheet.cell(row=3,column=1).value = report_period.date()
        sheet.cell(row=3,column=1).number_format = "yyyy年m月"
        sheet.cell(row=5,column=7).value = work_report.project_name
        sheet.cell(row=6,column=7).value = user_name

        count = 0
        for work_detail in work_details.values():
            sheet.cell(row=12+count,column=5).value = time_util.convert_to_serial_time(work_detail.start_time)
            sheet.cell(row=12+count,column=5).number_format = "[H]:MM"
            sheet.cell(row=12+count,column=9).value = time_util.convert_to_serial_time(work_detail.end_time)
            sheet.cell(row=12+count,column=9).number_format = "[H]:MM"
            sheet.cell(row=12+count,column=13).value = time_util.convert_to_serial_time(work_detail.break1_time)
            sheet.cell(row=12+count,column=13).number_format = "[H]:MM"
            sheet.cell(row=12+count,column=16).value = time_util.convert_to_serial_time(work_detail.break2_time)
            sheet.cell(row=12+count,column=16).number_format = "[H]:MM"
            sheet.cell(row=12+count,column=34).value = work_detail.status.name
            sheet.cell(row=12+count,column=38).value = work_detail.remarks

            count += 1


        sheet.cell(row=44,column=19).value = time_util.convert_to_serial_time(time_util.convert_to_clock_time(work_report.site_work_time))
        sheet.cell(row=44,column=19).number_format = "[H]:MM"


        # ==========================================================================
        # シートの設定
        # ==========================================================================
        # 入力を許可するセル
        editable_cells = []
        editable_cells.append(sheet["A3:I3"])
        editable_cells.append(sheet["G5:X6"])
        editable_cells.append(sheet["E12:R42"])
        editable_cells.append(sheet["S44:X44"])
        editable_cells.append(sheet["AH12:AZ42"])

        editable_cells = list(flatten(editable_cells))

        for cell in editable_cells:
            cell.protection = Protection(locked=False)

        sheet.protection.selectLockedCells = True # ロックされたセルは選択不可
        sheet.protection.password = PASSWORD # シートパスワードの設定
        sheet.protection.enable()
        # ==========================================================================

        response = HttpResponse(content_type='application/vnd.ms-excel')

        utf8_file_name = email.utils.encode_rfc2231(file_name, charset='UTF-8')
        response['Content-Disposition'] = f'attachment; filename*={utf8_file_name}'

        wb.save(response)

        return response

    return HttpResponse(status=500)

